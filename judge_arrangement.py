import openpyxl

# 数据读入
wb = openpyxl.load_workbook("赛程初版.xlsx") # 赛程文件
ws = wb['工作表1 (2)'] # 赛程工作表
wb_dict_judge = openpyxl.load_workbook("裁判ID表.xlsx")
ws_dict_judge = wb_dict_judge["Sheet1"]

# 包含比赛和小组信息的单元格寻找
match_cell = []
group_cell = []
for i in range(1,25):
    for j in range(1,7):
        value_temp = ws.cell(row=i, column=j).value
        if(value_temp != None):
            if(value_temp.__contains__("VS")):
                match_cell.append([i,j])
            else:
                if(value_temp.__contains__("组")):
                    group_cell.append([i,j])


# 分组字典生成
dict = {}
for i in group_cell:
    dict[ws.cell(row=i[0], column=i[1]).fill.fgColor.indexed] = ws.cell(row=i[0], column=i[1]).value


# 比赛信息提取
match_part = []
match_judge = []
match_time = []
match_group = []

for i in match_cell:
    if(i[1]<=4):
        match_part.append(ws.cell(row=i[0], column=i[1]).value.split()[0])
        match_judge.append(ws.cell(row=i[0], column=i[1]).value.split()[1].split("-"))
        match_time.append(ws.cell(row=i[0], column=1).value.split("-")[0].split("："))
        match_group.append(dict[ws.cell(row=i[0], column=i[1]).fill.fgColor.indexed])

# 比赛裁判信息转换
dict_judge = {}
for i in range(2, 46):
    dict_judge[ws_dict_judge.cell(row=i, column=3).value]=ws_dict_judge.cell(row=i, column=2).value

for i in range(len(match_judge)):
    for j in range(2):
        if(match_judge[i][j] == ''):
            match_judge[i][j]='空缺'
        if(dict_judge.__contains__(match_judge[i][j])):
            match_judge[i][j]=dict_judge[match_judge[i][j]]

# 推送内容生成
print("甲级")
for i in range(len(match_part)):
    if(match_group[i].__contains__("甲")):
        print(match_time[i][0], ":", match_time[i][1], "-", (int(match_time[i][0])+2), ":", match_time[i][1], sep="", end=" ")
        print(match_part[i])
        print("主裁：", match_judge[i][0], sep="", end=" ")
        print("副裁：", match_judge[i][1], sep="")
print("乙级")     
for i in range(len(match_part)):
    if(match_group[i].__contains__("乙")):
        print(match_time[i][0], ":", match_time[i][1], "-", (int(match_time[i][0])+1), ":", match_time[i][1], sep="", end=" ")
        print(match_part[i])
        print("主裁：", match_judge[i][0], sep="", end=" ")
        print("副裁：", match_judge[i][1], sep="")
